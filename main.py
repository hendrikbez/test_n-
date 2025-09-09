from flask import Flask, request, jsonify, session, redirect, render_template, url_for
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
import os
import win32com.client
import pythoncom
from pathlib import Path
import time
import shutil
import threading
import json
from datetime import datetime, timedelta
import locale

app = Flask(__name__)
app.secret_key = 'my-private-register-app-2025'

# Excel file path
EXCEL_FILE_PATH = r"C:\AutoHotkey\Kerk\ninja\2025_Register_Naamlys.xlsm"
EXCEL_BACKUP_PATH = r"C:\AutoHotkey\Kerk\ninja\2025_Register_Naamlys_backup.xlsm"

if not os.path.exists(EXCEL_FILE_PATH):
    print(f"WARNING: Excel file not found at C:\AutoHotkey\Kerk\ninja\2025_Register_Naamlys.xlsm")
else:
    print(f"Excel file found at C:\AutoHotkey\Kerk\ninja\2025_Register_Naamlys.xlsm")



# User credentials
USERS = {
    'admin': generate_password_hash('6371')
   
}

# Data storage
DATA_FILE = 'church_register.json'



def get_afrikaans_date():
    # Dictionary for Afrikaans month names
    afrikaans_months = {
        1: "Januarie", 2: "Februarie", 3: "Maart", 4: "April", 
        5: "Mei", 6: "Junie", 7: "Julie", 8: "Augustus", 
        9: "September", 10: "Oktober", 11: "November", 12: "Desember"
    }
    
    today = datetime.now()
    day = today.day
    month = afrikaans_months[today.month]
    year = today.year
    
    # Format as "05 September 2025"
    return f"{day:02d} {month} {year}"

# Function to update the Voorblad date
def update_voorblad_date():
    print("Starting update_voorblad_date()")
    try:
        # Initialize COM
        pythoncom.CoInitialize()
        
        # Connect to Excel
        excel = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Open the workbook
            wb = excel.Workbooks.Open(EXCEL_FILE_PATH)
            
            # Get the Voorblad sheet (Sheet4)
            try:
                sheet = wb.Worksheets('Voorblad')
            except:
                sheet = wb.Worksheets('Sheet4')  # Fallback if sheet name is different
            
            # Get formatted date
            afrikaans_date = get_afrikaans_date()
            
            # Update cells G14:N16 with the date
            # Merge the cells first if they aren't already merged
            try:
                sheet.Range("G14:N16").MergeCells = True
            except:
                pass  # Already merged
                
            # Set the value and formatting
            sheet.Range("G14:N16").Value = afrikaans_date
            
            # Updated formatting:
            sheet.Range("G14:N16").HorizontalAlignment = -4131  # xlLeft (-4131)
            sheet.Range("G14:N16").VerticalAlignment = -4108  # xlCenter
            sheet.Range("G14:N16").Font.Name = "Cambria"
            sheet.Range("G14:N16").Font.Size = 28
            sheet.Range("G14:N16").Font.Bold = True
            
            # Save the workbook
            wb.Save()
            wb.Close()
            excel.Quit()
            
            # Clean up COM
            pythoncom.CoUninitialize()
            
            return True
        except Exception as e:
            print(f"Error updating Voorblad date: {e}")
            if excel:
                try:
                    wb.Close(False)  # Close without saving
                    excel.Quit()
                except:
                    pass
            return False
    except Exception as e:
        print(f"COM initialization error: {e}")
        return False
    finally:
        try:
            pythoncom.CoUninitialize()
        except:
            pass


def excel_date_to_string(excel_date):
    """Convert Excel date number to string in dd-mmm format"""
    try:
        # Excel dates are number of days since 1900-01-01
        dt = datetime(1900, 1, 1) + timedelta(days=excel_date - 2)
        return dt.strftime('%d-%b')
    except:
        return str(excel_date)

def get_data_via_com():
    """Get data using Excel COM interface to avoid VBA corruption"""
    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        wb = excel.Workbooks.Open(EXCEL_FILE_PATH)
        sheet = wb.Worksheets('Register')
        
        data = []
        current_family_id = None
        
        # Find the last row with data
        last_row = sheet.Cells(sheet.Rows.Count, 2).End(-4162).Row  # -4162 = xlUp
        
        # Start from row 11 instead of row 2
        start_row = 11
        
        for row in range(start_row, last_row + 1):
            # Skip rows that are completely empty or are headers
            if sheet.Cells(row, 2).Value is None:
                continue
                
            # Get family ID (column A)
            family_id = sheet.Cells(row, 1).Value
            if family_id is not None and family_id != "":
                current_family_id = family_id
                
            # Get number (column B)
            number = sheet.Cells(row, 2).Value
            if number is not None:
                # Format dates properly
                verj = sheet.Cells(row, 5).Value
                huwelik = sheet.Cells(row, 6).Value
                
                # Convert Excel dates to proper format
                if verj and isinstance(verj, float):
                    try:
                        verj = excel_date_to_string(verj)
                    except:
                        verj = str(verj)
                
                if huwelik and isinstance(huwelik, float):
                    try:
                        huwelik = excel_date_to_string(huwelik)
                    except:
                        huwelik = str(huwelik)
                
                data.append({
                    'familyId': current_family_id,
                    'number': number,
                    'van': sheet.Cells(row, 3).Value or '',
                    'naam': sheet.Cells(row, 4).Value or '',
                    'verj': verj or '',
                    'huwelik': huwelik or '',
                    'selfoon': sheet.Cells(row, 7).Value or '',
                    'adres': sheet.Cells(row, 8).Value or '',
                    'epos': sheet.Cells(row, 9).Value or ''  # Added email column
                })
        
        wb.Close(False)  # Close without saving
        excel.Quit()
        return jsonify({'data': data})
    except Exception as e:
        if excel:
            try:
                excel.Quit()
            except:
                pass
        pythoncom.CoUninitialize()
        raise e

def get_data_via_openpyxl():
    """Fallback to openpyxl if COM fails"""
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True, data_only=True)
    sheet = wb['Register']
    
    data = []
    current_family_id = None
    
    # Start from row 11
    start_row = 11
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
        # Skip completely empty rows or header rows
        if all(cell is None for cell in row) or (row[1] is not None and isinstance(row[1], str) and row[1].upper() in ["NAAM", "NOMER", "NOMMER", "NUMBER"]):
            continue
            
        # Get family ID (column A) - use previous if empty
        family_id = row[0]
        if family_id is not None and family_id != "":
            current_family_id = family_id
            
        # Only add rows with data in column B (number)
        if row[1] is not None and not (isinstance(row[1], str) and row[1].upper() in ["NAAM", "NOMER", "NOMMER", "NUMBER"]):
            # Format dates properly
            verj = row[4]
            huwelik = row[5]
            
            # Convert Excel dates to proper format
            if verj and isinstance(verj, datetime):
                verj = verj.strftime('%d-%b')
            elif verj and isinstance(verj, str) and len(verj) > 10:
                # Try to parse date strings
                try:
                    dt = datetime.strptime(verj[:10], '%Y-%m-%d')
                    verj = dt.strftime('%d-%b')
                except:
                    pass
            
            if huwelik and isinstance(huwelik, datetime):
                huwelik = huwelik.strftime('%d-%b')
            elif huwelik and isinstance(huwelik, str) and len(huwelik) > 10:
                # Try to parse date strings
                try:
                    dt = datetime.strptime(huwelik[:10], '%Y-%m-%d')
                    huwelik = dt.strftime('%d-%b')
                except:
                    pass
            
            data.append({
                'familyId': current_family_id,
                'number': row[1],
                'van': row[2] or '',
                'naam': row[3] or '',
                'verj': verj or '',
                'huwelik': huwelik or '',
                'selfoon': row[6] or '',
                'adres': row[7] or '',
                'epos': row[8] or '' if len(row) > 8 else ''  # Added email column
            })
    
    wb.close()
    return jsonify({'data': data})

def add_data_operation(sheet, data):
    """Operation to add data via COM"""
    try:
        last_row = sheet.Cells(sheet.Rows.Count, 2).End(-4162).Row
        next_row = last_row + 1
        
        # Write data to columns C-I only, leaving A-B for macros
        sheet.Cells(next_row, 3).Value = data['van']
        sheet.Cells(next_row, 4).Value = data['naam']
        sheet.Cells(next_row, 5).Value = data['verj']
        sheet.Cells(next_row, 6).Value = data['huwelik']
        sheet.Cells(next_row, 7).Value = data['selfoon']
        sheet.Cells(next_row, 8).Value = data['adres']
        sheet.Cells(next_row, 9).Value = data.get('epos', '')  # Added email column
        
        # Return success but don't set a number since we're not setting column B
        return True, None
    except Exception as e:
        print(f"Add operation error: {e}")
        return False, None

def add_family_operation(sheet, data):
    """Operation to add family via COM"""
    try:
        last_row = sheet.Cells(sheet.Rows.Count, 2).End(-4162).Row
        next_row = last_row + 1
        
        members = data['members']
        family_adres = data.get('familyAdres', '')
        
        # Write each family member (only columns C-I)
        for i, member in enumerate(members):
            row_num = next_row + i
            
            # Don't set family ID or number (columns A-B)
            sheet.Cells(row_num, 3).Value = member['van']
            sheet.Cells(row_num, 4).Value = member['naam']
            sheet.Cells(row_num, 5).Value = member['verj']
            
            # Only include wedding date for the first member
            if i == 0:
                sheet.Cells(row_num, 6).Value = member.get('huwelik', '')
            else:
                sheet.Cells(row_num, 6).Value = ""
            
            sheet.Cells(row_num, 7).Value = member['selfoon']
            sheet.Cells(row_num, 8).Value = family_adres
            sheet.Cells(row_num, 9).Value = member.get('epos', '')  # Added email column
        
        # Return success but don't set a family ID since we're not setting column A
        return True, None
    except Exception as e:
        print(f"Add family operation error: {e}")
        return False, None

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {
        'register': [],  # Main member data
        'next_id': 1
    }

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)



@app.route('/test_excel', methods=['GET'])
def test_excel():
    """Test endpoint to check if Excel automation is working"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
    try:
        pythoncom.CoInitialize()
        excel = None
        try:
            # Try to create Excel application
            excel = win32com.client.dynamic.Dispatch("Excel.Application")
            version = excel.Version
            
            # Try to create a new workbook
            wb = excel.Workbooks.Add()
            sheet = wb.ActiveSheet
            sheet.Cells(1, 1).Value = "Test"
            
            # Close without saving
            wb.Close(False)
            excel.Quit()
            
            pythoncom.CoUninitialize()
            return jsonify({
                'success': True, 
                'message': f'Excel automation is working. Excel version: {version}'
            })
        except Exception as e:
            if excel:
                try:
                    excel.Quit()
                except:
                    pass
            pythoncom.CoUninitialize()
            return jsonify({'success': False, 'message': f'Excel automation test failed: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'COM initialization failed: {str(e)}'})

def create_register_pdf_python(excel_path):
    """Python implementation of CreateRegisterPDF macro"""
    try:
        # Preserve Sheet6 data
        sheet6_data = preserve_sheet6_data(excel_path)
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        if 'Register' not in wb.sheetnames:
            wb.close()
            return False, "Register sheet not found"
        
        register_sheet = wb['Register']
        
        # Create a new temporary workbook for the PDF
        from openpyxl import Workbook
        temp_wb = Workbook()
        temp_sheet = temp_wb.active
        temp_sheet.title = "Register"
        
        # Set up headers
        temp_sheet['A1'] = "Naam"
        temp_sheet['B1'] = "Van"
        temp_sheet['C1'] = "Telefoon #"
        
        # Set column widths
        temp_sheet.column_dimensions['A'].width = 15
        temp_sheet.column_dimensions['B'].width = 20
        temp_sheet.column_dimensions['C'].width = 15
        
        # Find the last row with data
        last_row = 10
        for row in range(11, register_sheet.max_row + 1):
            if register_sheet.cell(row=row, column=3).value is not None:
                last_row = row
        
        # Copy data from Register to temp sheet
        row_count = 1  # Start at row 1 (header)
        for i in range(11, last_row + 1):
            row_count += 1
            # Copy Name (column D in Register to column A in temp)
            temp_sheet.cell(row=row_count, column=1).value = register_sheet.cell(row=i, column=4).value
            # Copy Surname (column C in Register to column B in temp)
            temp_sheet.cell(row=row_count, column=2).value = register_sheet.cell(row=i, column=3).value
            # Copy Phone (column G in Register to column C in temp)
            temp_sheet.cell(row=row_count, column=3).value = register_sheet.cell(row=i, column=7).value
        
        # Apply formatting
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        # Format header
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        for col in range(1, 4):
            cell = temp_sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Apply alternating row colors
        for i in range(2, row_count + 1):
            fill_color = "F2F2F2" if i % 2 == 0 else "FFFFE0"
            for col in range(1, 4):
                cell = temp_sheet.cell(row=i, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Apply borders
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row in range(1, row_count + 1):
            for col in range(1, 4):
                temp_sheet.cell(row=row, column=col).border = thin_border
        
        # Save the temporary workbook
        import os
        pdf_dir = r"C:\Drukwerk\Lys"
        if not os.path.exists(pdf_dir):
            os.makedirs(pdf_dir)
        
        temp_file = os.path.join(pdf_dir, "Register_Temp.xlsx")
        temp_wb.save(temp_file)
        
        # Convert to PDF using a library like pywin32 or a PDF converter
        # For this example, we'll just save the Excel file
        pdf_path = os.path.join(pdf_dir, "Register.pdf")
        
        # Here you would use a PDF conversion library
        # For now, we'll just indicate success
        
        # Clean up
        temp_wb.close()
        wb.close()
        
        # Restore Sheet6 data
        if sheet6_data:
            restore_sheet6_data(excel_path, sheet6_data)
        
        return True, f"Register Excel file created at {temp_file}. PDF conversion requires additional setup."
    
    except Exception as e:
        print(f"Error creating register PDF: {e}")
        return False, str(e)

def print_to_pdf_landscape_python(excel_path):
    """Python implementation of PrintToPDF_Landscape1 macro"""
    try:
        # Preserve Sheet6 data
        sheet6_data = preserve_sheet6_data(excel_path)
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        if 'Register' not in wb.sheetnames:
            wb.close()
            return False, "Register sheet not found"
        
        # Define the output path
        import os
        pdf_dir = r"C:\Drukwerk\Lys"
        if not os.path.exists(pdf_dir):
            os.makedirs(pdf_dir)
        
        pdf_path = os.path.join(pdf_dir, "Lys.pdf")
        
        # For now, we'll just save a copy of the Excel file
        # In a real implementation, you would use a PDF conversion library
        excel_copy = os.path.join(pdf_dir, "Lys.xlsx")
        wb.save(excel_copy)
        wb.close()
        
        # Restore Sheet6 data
        if sheet6_data:
            restore_sheet6_data(excel_path, sheet6_data)
        
        return True, f"Excel file saved at {excel_copy}. PDF conversion requires additional setup."
    
    except Exception as e:
        print(f"Error printing to PDF: {e}")
        return False, str(e)


def preserve_sheet6_data(excel_path):
    """Read and preserve data from Sheet6 (Start)"""
    try:
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        if 'Start' in wb.sheetnames:
            sheet = wb['Start']
            # Store all data from the sheet
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            wb.close()
            return data
        else:
            wb.close()
            return None
    except Exception as e:
        print(f"Error preserving Sheet6 data: {e}")
        return None

def restore_sheet6_data(excel_path, data):
    """Restore data to Sheet6 (Start)"""
    if not data:
        return False
        
    try:
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        if 'Start' in wb.sheetnames:
            sheet = wb['Start']
            # Clear existing data
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None
            
            # Restore data
            for i, row_data in enumerate(data, 1):
                for j, value in enumerate(row_data, 1):
                    sheet.cell(row=i, column=j).value = value
            
            wb.save(excel_path)
            wb.close()
            return True
        else:
            wb.close()
            return False
    except Exception as e:
        print(f"Error restoring Sheet6 data: {e}")
        return False


def update_voorblad_date_with_openpyxl(excel_path):
    """Update the date in Sheet4 (Voorblad) merged cells G14:N16 using openpyxl"""
    try:
        # Get formatted date
        afrikaans_date = get_afrikaans_date()
        
        # Open the workbook
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        # Get the Voorblad sheet
        sheet_name = 'Voorblad' if 'Voorblad' in wb.sheetnames else 'Sheet4'
        sheet = wb[sheet_name]
        
        # First, unmerge the cells if they are merged
        try:
            sheet.unmerge_cells('G14:N16')
        except:
            pass  # Cells might not be merged
        
        # Update cell G14 with the date
        sheet['G14'] = afrikaans_date
        
        # Apply formatting
        from openpyxl.styles import Font, Alignment
        cell = sheet['G14']
        cell.font = Font(name='Cambria', size=28, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Merge the cells again
        sheet.merge_cells('G14:N16')
        
        # Save the workbook
        wb.save(excel_path)
        wb.close()
        
        return True, f"Datum opgedateer na {afrikaans_date}"
    except Exception as e:
        print(f"Error updating Voorblad date with openpyxl: {e}")
        return False, str(e)

def update_voorblad_date_simple(excel_path):
    """A simpler approach to update the date in Voorblad"""
    try:
        # Get formatted date
        afrikaans_date = get_afrikaans_date()
        
        # Make a backup of the file first
        backup_path = excel_path.replace('.xlsm', '_backup.xlsm')
        import shutil
        shutil.copy2(excel_path, backup_path)
        
        # Use openpyxl with minimal operations
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        # Get the Voorblad sheet
        sheet_name = None
        for name in wb.sheetnames:
            if name.lower() in ['voorblad', 'sheet4']:
                sheet_name = name
                break
                
        if not sheet_name:
            wb.close()
            return False, "Voorblad sheet not found"
            
        sheet = wb[sheet_name]
        
        # Simply update cell G14 with the date
        sheet['G14'] = afrikaans_date
        
        # Save and close
        wb.save(excel_path)
        wb.close()
        
        print(f"Date updated to {afrikaans_date}")
        return True, f"Datum opgedateer na {afrikaans_date}"
        
    except Exception as e:
        print(f"Error in update_voorblad_date_simple: {e}")
        # Try to restore from backup if we made one
        try:
            if 'backup_path' in locals():
                shutil.copy2(backup_path, excel_path)
                print("Restored from backup after error")
        except:
            pass
        return False, str(e)



def sort_format_and_number_in_python(excel_path):
    """
    Python implementation of the SortFamilyMembers macro functionality.
    """
    try:
        # Load the workbook with openpyxl
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        # Get the Register sheet
        if 'Register' not in wb.sheetnames:
            return False, "Register sheet not found"
        
        sheet = wb['Register']
        
        # Find the last row with data
        last_row = 10  # Start from row 11 (index 10)
        for row in range(11, sheet.max_row + 1):
            if sheet.cell(row=row, column=3).value is not None:  # Check column C (Van)
                last_row = row
        
        if last_row <= 10:
            return False, "No data found to sort"
        
        # Read all data into a list
        data = []
        for row in range(11, last_row + 1):
            if sheet.cell(row=row, column=3).value is not None:  # Only include rows with data in Van column
                data.append({
                    'row': row,
                    'van': sheet.cell(row=row, column=3).value or '',
                    'naam': sheet.cell(row=row, column=4).value or '',
                    'verj': sheet.cell(row=row, column=5).value,
                    'huwelik': sheet.cell(row=row, column=6).value,
                    'selfoon': sheet.cell(row=row, column=7).value or '',
                    'adres': sheet.cell(row=row, column=8).value or '',
                    'epos': sheet.cell(row=row, column=9).value or ''
                })
        
        # Sort the data by Van and then by Naam
        data.sort(key=lambda x: (str(x['van']).lower(), str(x['naam']).lower()))
        
        # Assign family IDs and sequential numbers
        current_family = None
        family_id = 0
        
        for i, item in enumerate(data):
            # Assign sequential number
            seq_num = i + 1
            
            # Assign family ID
            if item['van'] != current_family:
                family_id += 1
                current_family = item['van']
            
            # Update the item with IDs
            item['family_id'] = family_id
            item['seq_num'] = seq_num
        
        # Write the sorted data back to the sheet
        for i, item in enumerate(data):
            row = 11 + i
            
            # Write family ID and sequence number
            sheet.cell(row=row, column=1).value = item['family_id']
            sheet.cell(row=row, column=2).value = item['seq_num']
            
            # Write the rest of the data
            sheet.cell(row=row, column=3).value = item['van']
            sheet.cell(row=row, column=4).value = item['naam']
            sheet.cell(row=row, column=5).value = item['verj']
            sheet.cell(row=row, column=6).value = item['huwelik']
            sheet.cell(row=row, column=7).value = item['selfoon']
            sheet.cell(row=row, column=8).value = item['adres']
            sheet.cell(row=row, column=9).value = item['epos']
        
        # Save the workbook
        wb.save(excel_path)
        wb.close()
        
        return True, "Families suksesvol gesorteer en genommer"
        
    except Exception as e:
        print(f"Error in sort_format_and_number_in_python: {e}")
        return False, f"Error: {str(e)}"






@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username in USERS and check_password_hash(USERS[username], password):
            session['user'] = username
            
            # Update the Voorblad date when user logs in
            try:
                update_voorblad_date()
            except Exception as e:
                print(f"Failed to update Voorblad date: {e}")
                # Continue even if date update fails
                
            return jsonify({'success': True})
        return jsonify({'success': False, 'message': 'Ongeldige aanmelding'})
    
    return render_template('login.html')


@app.route('/add_person', methods=['POST'])
def add_person():
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401
    
    try:
        data = request.json
        
        if not data.get('van') or not data.get('naam'):
            return jsonify({'success': False, 'message': 'Van en naam is verpligtend'}), 400
        
        # Format phone number if provided
        selfoon = data.get('selfoon', '')
        if selfoon:
            digits = ''.join(filter(str.isdigit, selfoon))
            if len(digits) >= 10:
                selfoon = f"{digits[:3]} {digits[3:6]} {digits[6:10]}"
        
        try:
            # Load the workbook with openpyxl
            # keep_vba=True preserves any VBA code in the file
            wb = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True)
            
            # Get the Register sheet
            if 'Register' in wb.sheetnames:
                sheet = wb['Register']
            else:
                return jsonify({'success': False, 'message': 'Register sheet not found in Excel file'})
            
            # Find the last row with data in column C
            last_row = 10  # Start from row 11 (index 10)
            for row in range(11, sheet.max_row + 1):
                if sheet.cell(row=row, column=3).value is not None:
                    last_row = row
            
            next_row = last_row + 1
            
            # Write data to the new row
            sheet.cell(row=next_row, column=3).value = data['van']
            sheet.cell(row=next_row, column=4).value = data['naam']
            sheet.cell(row=next_row, column=5).value = data.get('verj', '')
            sheet.cell(row=next_row, column=6).value = data.get('huwelik', '')
            sheet.cell(row=next_row, column=7).value = selfoon
            sheet.cell(row=next_row, column=8).value = data.get('adres', '')
            sheet.cell(row=next_row, column=9).value = data.get('epos', '')
            
            # Save the workbook
            wb.save(EXCEL_FILE_PATH)
            wb.close()
            
            return jsonify({'success': True, 'message': 'Persoon suksesvol bygevoeg'})
            
        except Exception as e:
            print(f"Error adding person with openpyxl: {e}")
            return jsonify({'success': False, 'message': f"Could not add person: {str(e)}"})
            
    except Exception as e:
        print(f"Error in add_person: {e}")
        return jsonify({'success': False, 'message': str(e)})





@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/dashboard')
@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect(url_for('index'))
    return render_template('dashboard.html')



@app.route('/run_macro', methods=['POST'])
def run_macro_endpoint():  # Changed function name to avoid conflict
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401
    
    data = request.json
    macro_name = data.get('macroName')
    
    if not macro_name:
        return jsonify({'success': False, 'message': 'Geen makro naam verskaf nie'}), 400

    # Preserve Sheet6 data before running any macro
    sheet6_data = preserve_sheet6_data(EXCEL_FILE_PATH)
    
    try:
        # Choose the appropriate implementation based on the macro name
        if macro_name == 'SortFamilyMembers':
            success, message = sort_format_and_number_in_python(EXCEL_FILE_PATH)
        elif macro_name == 'CreateRegisterPDF':
            success, message = create_register_pdf_python(EXCEL_FILE_PATH)
        elif macro_name == 'PrintToPDF_Landscape1':
            success, message = print_to_pdf_landscape_python(EXCEL_FILE_PATH)
        elif macro_name == 'UpdateDate':
            success, message = update_voorblad_date_with_openpyxl(EXCEL_FILE_PATH)
        else:
            # For other macros, try the VBS approach
            try:
                success, message = run_macro_via_vbs(EXCEL_FILE_PATH, macro_name)
            except Exception as e:
                success = False
                message = f"VBS execution failed: {str(e)}"
        
        # Restore Sheet6 data if needed
        if sheet6_data:
            restore_sheet6_data(EXCEL_FILE_PATH, sheet6_data)
        
        return jsonify({'success': success, 'message': message})
    
    except Exception as e:
        print(f"Error running macro '{macro_name}': {e}")
        
        # Attempt to restore Sheet6 data even if there was an error
        if sheet6_data:
            restore_sheet6_data(EXCEL_FILE_PATH, sheet6_data)
        
        return jsonify({'success': False, 'message': f"Error: {str(e)}"})






def run_macro_via_vbs(excel_path, macro_name):
    """Run an Excel macro using a VBS script"""
    try:
        # Create a temporary VBS script
        vbs_path = os.path.join(os.path.dirname(excel_path), "run_macro_temp.vbs")
        
        # Map macro names to their full paths if needed
        macro_mapping = {
            'SortFamilyMembers': 'Sheet3.SortFamilyMembers',
            'CreateRegisterPDF': 'Sheet3.CreateRegisterPDF',
            'PrintToPDF_Landscape': 'Sheet3.PrintToPDF_Landscape1',
            'UpdateDate': 'ThisWorkbook.update_voorblad_date'
        }
        
        full_macro_name = macro_mapping.get(macro_name, macro_name)
        
        # Write the VBS script
        with open(vbs_path, 'w') as f:
            f.write(f'''
            On Error Resume Next
            
            ' Display any errors that occur
            Sub DisplayError()
                If Err.Number <> 0 Then
                    WScript.Echo "Error " & Err.Number & ": " & Err.Description
                    WScript.Quit 1
                End If
            End Sub
            
            ' Run the macro
            Sub RunMacro()
                Dim xl, wb
                
                ' Create Excel application
                Set xl = CreateObject("Excel.Application")
                If Err.Number <> 0 Then
                    WScript.Echo "Failed to create Excel application: " & Err.Description
                    WScript.Quit 1
                End If
                
                ' Make Excel visible and disable alerts
                xl.Visible = True
                xl.DisplayAlerts = False
                
                ' Open the workbook
                Set wb = xl.Workbooks.Open("{excel_path}")
                If Err.Number <> 0 Then
                    WScript.Echo "Failed to open workbook: " & Err.Description
                    xl.Quit
                    WScript.Quit 1
                End If
                
                ' Run the macro
                xl.Run "{full_macro_name}"
                If Err.Number <> 0 Then
                    WScript.Echo "Failed to run macro: " & Err.Description
                    wb.Close False
                    xl.Quit
                    WScript.Quit 1
                End If
                
                ' Save and close
                wb.Save
                wb.Close
                xl.Quit
                
                WScript.Echo "Success"
            End Sub
            
            ' Main execution
            RunMacro
            DisplayError
            ''')
        
        # Run the VBS script
        import subprocess
        result = subprocess.run(['cscript', '//NoLogo', vbs_path], capture_output=True, text=True)
        
        # Clean up
        try:
            os.remove(vbs_path)
        except:
            pass
        
        # Check the result
        if result.returncode == 0 and "Success" in result.stdout:
            return True, f"Makro '{macro_name}' suksesvol uitgevoer"
        else:
            error_msg = result.stdout.strip() or result.stderr.strip() or f"Unknown error running macro '{macro_name}'"
            print(f"VBS Error: {error_msg}")
            return False, error_msg
            
    except Exception as e:
        print(f"Error in run_macro_via_vbs: {e}")
        return False, str(e)




@app.route('/get_current_date')
def get_current_date():
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401
    
    try:
        # Initialize COM
        pythoncom.CoInitialize()
        
        # Connect to Excel
        excel = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Open the workbook
            wb = excel.Workbooks.Open(EXCEL_FILE_PATH)
            
            # Get the Voorblad sheet (Sheet4)
            try:
                sheet = wb.Worksheets('Voorblad')
            except:
                sheet = wb.Worksheets('Sheet4')  # Fallback if sheet name is different
            
            # Get the current date from the sheet
            current_date = sheet.Range("G14").Value
            
            wb.Close(False)  # Close without saving
            excel.Quit()
            
            # Clean up COM
            pythoncom.CoUninitialize()
            
            return jsonify({'success': True, 'date': current_date or get_afrikaans_date()})
        except Exception as e:
            print(f"Error getting date: {e}")
            if excel:
                try:
                    wb.Close(False)  # Close without saving
                    excel.Quit()
                except:
                    pass
            return jsonify({'success': False, 'message': str(e)})
    except Exception as e:
        print(f"COM initialization error: {e}")
        return jsonify({'success': False, 'message': str(e)})
    finally:
        try:
            pythoncom.CoUninitialize()
        except:
            pass

@app.route('/update_date', methods=['POST'])
def update_date():
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401
    
    try:
        print("Updating date...")
        # Use the simpler function
        success, message = update_voorblad_date_simple(EXCEL_FILE_PATH)
        
        if success:
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message})
    except Exception as e:
        print(f"Error in update_date endpoint: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/recover_excel', methods=['GET'])
def recover_excel():
    """Emergency endpoint to restore Excel file from backup"""
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
    try:
        backup_path = EXCEL_FILE_PATH.replace('.xlsm', '_backup.xlsm')
        
        # Check if backup exists
        if not os.path.exists(backup_path):
            return jsonify({'success': False, 'message': 'No backup file found'})
            
        # Restore from backup
        import shutil
        shutil.copy2(backup_path, EXCEL_FILE_PATH)
        
        return jsonify({'success': True, 'message': 'Excel file restored from backup'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/get_data')
def get_data():
    if 'user' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        # First try to use COM for reading to preserve VBA integrity
        try:
            return get_data_via_com()
        except Exception as com_error:
            print(f"COM read failed, trying openpyxl: {com_error}")
            return get_data_via_openpyxl()
    except Exception as e:
        print(f"Error loading data: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/SortFamilyMembers', methods=['POST'])
def sort_family_members():
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Not authenticated'}), 401
    
    try:
        # Initialize COM
        pythoncom.CoInitialize()
        
        # Connect to Excel
        excel = None
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            # Open the workbook
            wb = excel.Workbooks.Open(EXCEL_FILE_PATH)
            
            # Get the Register sheet
            sheet = wb.Worksheets('Register')
            
            # Find the last row with data
            last_row = sheet.Cells(sheet.Rows.Count, 2).End(-4162).Row
            
            # Define the range to sort (from row 11 to last_row, columns A to I)
            sort_range = sheet.Range(f"A11:I{last_row}")
            
            # Sort by Van (column C) and then by Naam (column D)
            sort_range.Sort(
                Key1=sheet.Range("C11"),
                Order1=1,  # xlAscending
                Key2=sheet.Range("D11"),
                Order2=1,  # xlAscending
                Header=1  # xlYes (include headers)
            )
            
            # Save the workbook
            wb.Save()
            wb.Close()
            excel.Quit()
            
            # Clean up COM
            pythoncom.CoUninitialize()
            
            return jsonify({'success': True, 'message': 'Register gesorteer volgens familienaam'})
        except Exception as e:
            print(f"Error sorting register: {e}")
            if excel:
                try:
                    wb.Close(False)  # Close without saving
                    excel.Quit()
                except:
                    pass
            return jsonify({'success': False, 'message': str(e)})
    except Exception as e:
        print(f"COM initialization error: {e}")
        return jsonify({'success': False, 'message': str(e)})
    finally:
        try:
            pythoncom.CoUninitialize()
        except:
            pass



@app.route('/some_long_operation')
def long_operation():
    try:
        # Do some work
        session['progress'] = 25
        # More work
        session['progress'] = 50
        # Final work
        session['progress'] = 100
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})






 # Add this to your main.py after the imports
@app.before_request
def before_request():
    if 'progress' not in session:
        session['progress'] = 0

@app.after_request
def after_request(response):
    # Reset progress after each request
    session['progress'] = 0
    return response

@app.route('/api/progress')
def get_progress():
    return jsonify({'progress': session.get('progress', 0)})





@app.route('/register')
def register_view():
    if 'user' not in session:
        return redirect(url_for('index'))
    return render_template('register.html')

if __name__ == '__main__':
    print("Starting Flask server on http://127.0.0.1:5000")
    app.run(debug=True, port=5000)
